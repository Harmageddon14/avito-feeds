# -*- coding: utf-8 -*-
"""
Сборка кампании 2026-07-21 в канон-фид promles.xlsx.

Отличие от build_campaign.py: НЕ чистит лист. Существующие живые строки
остаются нетронутыми, новые объявления ДОПИСЫВАЮТСЯ в конец.

usage:
  python3 assemble.py            # линт + дедуп, без записи (сухой прогон)
  python3 assemble.py --build    # + уникализация фото, push на GitHub, запись в фид
"""
import os, sys, json, glob, re
sys.path.insert(0, os.path.expanduser("~/avito-tools"))

import openpyxl
import accounts, config, idgen, lint, photos as photomod, ghpush, photoclass

HERE = os.path.dirname(os.path.abspath(__file__))
CANON = os.path.expanduser("~/avito-feeds/promles.xlsx")
ACCOUNT = "438889381"
CAMP = "kedr-sosna-lezhnevka-2026-07-21"


def load():
    spec = json.load(open(os.path.join(HERE, "spec.json"), encoding="utf-8"))
    texts = {}
    for f in sorted(glob.glob(os.path.join(HERE, "texts_*.json"))):
        texts.update(json.load(open(f, encoding="utf-8")))
    missing = [a["code"] for a in spec["ads"] if a["code"] not in texts]
    if missing:
        print(f"✗ нет текстов для: {', '.join(missing)}")
        sys.exit(1)
    for a in spec["ads"]:
        a["title"] = texts[a["code"]]["title"]
        a["description"] = texts[a["code"]]["description"]
    return spec


def run_lint(spec):
    errors = 0
    print("=== ЛИНТ ===")
    for a in spec["ads"]:
        issues = lint.lint_title(a["title"]) + lint.lint_description(a["description"])
        hard = [i for i in issues if i[0] == "error"]
        if issues:
            print(f"[{a['code']}] {a['title']}")
            for sev, msg in issues:
                print(f"    {'X' if sev=='error' else '.'} {sev.upper():5} {msg}")
        errors += len(hard)
    print(f"ошибок: {errors}")
    return errors


def dup_check(spec):
    """Кросс-дедуп: заголовки уникальны, описания не повторяются по 6-граммам."""
    print("=== ДЕДУП ===")
    bad = 0
    titles = [a["title"].strip().lower() for a in spec["ads"]]
    ex = [t.strip().lower() for t in spec["existing_titles"]]
    for t in set(titles):
        if titles.count(t) > 1:
            print(f"    X дубль заголовка: {t}")
            bad += 1
    for t in titles:
        if t in ex:
            print(f"    X заголовок совпал с живым: {t}")
            bad += 1

    def grams(s):
        w = re.findall(r"[а-яёa-z]+", lint.strip_html(s).lower())
        return {" ".join(w[i:i+6]) for i in range(len(w) - 5)}

    g = {a["code"]: grams(a["description"]) for a in spec["ads"]}
    codes = [a["code"] for a in spec["ads"]]
    for i in range(len(codes)):
        for j in range(i + 1, len(codes)):
            a, b = g[codes[i]], g[codes[j]]
            if not a or not b:
                continue
            ov = len(a & b) / min(len(a), len(b))
            if ov > 0.12:
                print(f"    X пересечение {codes[i]}/{codes[j]}: {ov:.0%}")
                bad += 1
    print(f"проблем: {bad}")
    return bad


def photo_order(spec):
    print("=== ФОТО: живое первым ===")
    bad = 0
    for a in spec["ads"]:
        ordered, n_live = photoclass.order_live_first(a["photos"])
        a["photos"] = ordered
        if n_live == 0:
            print(f"    X {a['code']} нет живого фото")
            bad += 1
    print(f"без живого: {bad}")
    return bad


def build(spec):
    acc = accounts.get(ACCOUNT)
    defaults = config.content_defaults(acc)

    taken = idgen.existing_ids(CANON)
    for a in spec["ads"]:
        a["id"] = idgen.new_id(taken)

    work = os.path.expanduser(f"~/avito-tools/build/{ACCOUNT}/{CAMP}")
    seen = set()
    n = 0
    for i, a in enumerate(spec["ads"]):
        outs = photomod.uniquify(a["photos"], work, ad_no=i + 1, seen_md5=seen)
        a["image_urls"] = ghpush.push_images(outs, f"{ACCOUNT}/{CAMP}/{a['id']}")
        n += len(outs)
        print(f"  фото {a['code']} -> {a['id']} ({len(outs)})")
    print(f"=== ФОТО: уникализировано и залито {n} шт ===")

    wb = openpyxl.load_workbook(CANON)
    ws = wb[config.TOVAR_SHEET]
    cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v:
            cols[str(v).strip()] = c
    idc = cols["Уникальный идентификатор объявления"]
    row = 5
    while ws.cell(row, idc).value:
        row += 1
    print(f"=== ЗАПИСЬ === существующих строк: {row-5}, дописываем с строки {row}")

    def setv(r, name, val):
        if name in cols and val is not None:
            ws.cell(r, cols[name]).value = val

    from datetime import datetime, timedelta
    for a in spec["ads"]:
        for k, v in defaults.items():
            setv(row, k, v)
        setv(row, "Уникальный идентификатор объявления", a["id"])
        setv(row, "Название объявления", a["title"])
        setv(row, "Описание объявления", a["description"])
        setv(row, "Адрес", a["address"])
        setv(row, "Цена", a["price"])
        setv(row, "Вид древесины", a["species"])
        setv(row, "Длина", a["length"])
        setv(row, "Диаметр", a["diameter"])
        setv(row, "Ссылки на фото", " | ".join(a["image_urls"]))
        setv(row, "DateBegin", a["dt"])
        dt = datetime.strptime(a["dt"], "%Y-%m-%dT%H:%M:%S+03:00")
        setv(row, "AvitoDateEnd",
             (dt + timedelta(hours=config.AVITO_DATEEND_OFFSET_H)).strftime(config.DATEBEGIN_FMT))
        row += 1

    wb.save(CANON)
    total = row - 5
    print(f"OK канон-фид: {CANON}, всего объявлений: {total}")
    json.dump({a["code"]: {"id": a["id"], "title": a["title"], "city": a["city"],
                           "price": a["price"], "dt": a["dt"], "address": a["address"]}
               for a in spec["ads"]},
              open(os.path.join(HERE, "published.json"), "w"), ensure_ascii=False, indent=1)
    return total


if __name__ == "__main__":
    spec = load()
    print(f"объявлений в кампании: {len(spec['ads'])}")
    e = run_lint(spec)
    d = dup_check(spec)
    p = photo_order(spec)
    if e or d or p:
        print("\nX есть блокеры, сборка не выполнена")
        sys.exit(1)
    if "--build" in sys.argv:
        build(spec)
    else:
        print("\nсухой прогон чист. Запусти с --build для сборки.")
